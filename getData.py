import openpyxl
import pandas as pd

from basic import get_cookie, request1_2, request2, analyze_string

df = pd.read_excel("2023stu-all.xlsx")
wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1, column=1, value="ID")
ws.cell(row=1, column=2, value="学号")
ws.cell(row=1, column=3, value="姓名")

# [[NailFecMODIFY]]
selval = 229
subresult = request1_2(selval)
allsubsn = [item["exasubSN"] for item in subresult]

for iPerson, row in df.iterrows():
    iPerson = int(iPerson)
    cNum = row["num"]
    name = row["name"]
    sNum = row["cookieB"]
    print(cNum, name, sNum)
    text = get_cookie(sNum, cNum, name)

    nsub = len(allsubsn)
    result = request2(selval, allsubsn, text)

    ws.cell(row=iPerson + 2, column=1, value=sNum)
    ws.cell(row=iPerson + 2, column=2, value=cNum)
    ws.cell(row=iPerson + 2, column=3, value=name)

    allData = analyze_string(result, nsub, True)
    for i, data in enumerate(allData):
        ws.cell(row=iPerson + 2, column=i + 4, value=data)
    print(allData)

lastp = 0
for i in range(nsub):
    pa = result.find("</th><th>", lastp)
    pb = result.find("</th>", pa + 2)
    thisData = result[(pa + len("</th><th>")): pb]
    ws.cell(row=1, column=i + 4, value=thisData)
    ws.cell(row=1, column=i + 4 + nsub, value=thisData)
    lastp = pa + 1

wb.save("lsoutput.xlsx")  # [[NailFecMODIFY]]
