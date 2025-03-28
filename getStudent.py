import json

import openpyxl
import pandas as pd

from basic import get_cookie, request0, request1, request2, analyze_string

df = pd.read_excel("2023stu-all.xlsx")
wb = openpyxl.Workbook()
ws = wb.active

cNum = int(input("input cNum: "))
if cNum == 1:
    cNum = 728

row = df[df["num"] == cNum]
if not row.empty:
    name = row["name"].values[0]
    sNum = row["cookieB"].values[0]
else:
    print(f"cNum {cNum} not found")
    exit(0)

text = get_cookie(sNum, cNum, name)

result = request0(text)
result = json.loads(result)
result = result["rows"]

iExam = 0
for exam in result:
    iExam += 1
    sn = exam["sn"]
    name = exam["name"]
    print(sn, name)
    ws.cell(row=iExam, column=1, value=sn)
    ws.cell(row=iExam, column=2, value=name)

    result = request1(sn, text)
    result = json.loads(result)

    allsubname = []
    allsubsn = []
    for eexam in result:
        subSN = eexam["subSN"]
        subName = eexam["subName"]
        exasubSN = eexam["exasubSN"]
        exaSubIsExpr = eexam["exaSubIsExpr"]
        allsubname.append(subName)
        allsubsn.append(exasubSN)
    nsub = len(allsubsn)
    allsubsn = ",".join(allsubsn)
    # print(allsubsn)    # [[NailFecDEBUG]]

    result = request2(sn, allsubsn, text)

    allData = analyze_string(result, nsub, True)
    for i, data in enumerate(allsubname):
        ws.cell(row=iExam, column=i + 3, value=data)
    for i, data in enumerate(allData):
        ws.cell(row=iExam, column=i + len(allsubname) + 3, value=data)
    print(allsubname)
    print(allData)

wb.save("lsoutput.xlsx")
