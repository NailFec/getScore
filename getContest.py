# before using it, modify "NailSEARCHtomodify"
import base64
import gzip

import openpyxl
import pandas as pd
import requests

wb = openpyxl.Workbook()
ws = wb.active

for tryid in range(213, 1000):
	print(tryid)
	try:
		# Part 1: get data from Excel
		num = '728'
		name = '顾铭屹'
		cookieB = '255'
		index = 1

		# Part 2: turn cookieB to gzip
		def compress_and_encode(text):
			compressed_data = gzip.compress(text.encode('utf-8'))
			encoded_data = base64.b64encode(compressed_data)
			return encoded_data

		original_text = f"0※{cookieB}※3101062002320230{num}※{name}※0※"
		# print(original_text)
		compressed_and_encoded_data = compress_and_encode(original_text)
		# print(compressed_and_encoded_data.decode('utf-8'))

		# Part 3: get request content
		url = "https://semp.xinghuiyuan.com.cn/stutkYczx/ExamScore/Handlers/ExamScore.ashx?type=getsub&date=1716095218033"
		headers = {"Accept": "*/*", "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6,ja;q=0.5",
			"Cache-Control": "no-cache", "Connection": "keep-alive",
			"Cookie": f"ASP.NET_SessionId=vwtbtzvfkiq34fwbstshuhna; USER_INFO_STU={compressed_and_encoded_data.decode('utf-8')}",
			"Origin": "https://semp.xinghuiyuan.com.cn", "Pragma": "no-cache",
			"Referer": "https://semp.xinghuiyuan.com.cn/stutkYczx/ExamScore/StatAnalysis.aspx",
			"Sec-Fetch-Dest": "empty", "Sec-Fetch-Mode": "cors", "Sec-Fetch-Site": "same-origin",
			"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36 Edg/125.0.0.0",
			"Content-Type": "application/x-www-form-urlencoded", "DNT": "1",
			"sec-ch-ua": '"Microsoft Edge";v="125", "Chromium";v="125", "Not.A/Brand";v="24"', "sec-ch-ua-mobile": "?0",
			"sec-ch-ua-platform": '"Windows"', "sec-gpc": "1"}
		data = { "exalist": tryid }
		response = requests.post(url, headers=headers, data=data)
		# f.write(response.text)
		# f.write('\n')
		print(response.text) ##

		# Part 4: find out the useful info
		findstra = "</td><td>"
		findstrb = "</td>"
		lastp = 0
		ws.cell(row=index + 1, column=1, value=cookieB)
		ws.cell(row=index + 1, column=2, value=num)
		ws.cell(row=index + 1, column=3, value=name)
		for i in range(1, 11):  # n(allsubsn)*10+1 NailSEARCHtomodify
			pa = response.text.find(findstra, lastp)
			pb = response.text.find(findstrb, pa + 2)
			thissub = response.text[pa + len(findstra):pb]
			lastp = pa + 2
			# print(f"data{i}: {thissub}")
			ws.cell(row=index + 1, column=i + 3, value=thissub)
	except Exception as e:
		print(f"NailERROR on index {index}: {e}")

# time.sleep(0.1)

# wb.save("onlytest.xlsx")  # NailSEARCHtomodify
